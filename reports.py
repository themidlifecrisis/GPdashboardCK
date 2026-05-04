"""Accountant report builder — pure logic, no Streamlit or I/O.

Two public functions:
- build_month_report(month, cos_other_pct, budget_df, actuals_df)
    Returns a structured report dict.
- build_excel_workbook(report)
    Returns xlsx bytes (Summary sheet today; per-branch detail sheets
    are added in Task 6).
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from calculations import (
    actuals_from_row,
    calculate_gp_actuals,
    targets_from_budget,
)
from config import BRANCHES


# Visual constants
HEADER_FILL = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FONT = Font(bold=True)
SECTION_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
SECTION_FONT = Font(bold=True, color="1E88E5")

GREEN_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
RED_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

CURRENCY_FMT = "#,##0;(#,##0);-"
PCT_FMT = "0.0%"


def _empty_targets() -> dict:
    """Zero-valued target dict, same shape as calculate_targets output."""
    return {
        "sales_target": 0,
        "paint_labour_target": 0,
        "parts_sales_target": 0,
        "parts_markup": 0,
        "parts_costs": 0,
        "cos_other": 0,
        "rsb_paint": 0,
        "consumables": 0,
        "gross_profit": 0,
        "gp_pct": 0.0,
    }


def _row_for_branch(branch: str, targets: dict, actuals: dict, extras: dict) -> dict:
    """Flatten one branch's data into the Summary-row schema."""
    return {
        "branch": branch,
        "sales_target": targets["sales_target"],
        "sales_actual": actuals["sales"],
        "sales_var": actuals["sales"] - targets["sales_target"],
        "parts_costs_target": targets["parts_costs"],
        "parts_costs_actual": actuals["parts_costs"],
        "parts_costs_var": actuals["parts_costs"] - targets["parts_costs"],
        "cos_other_target": targets["cos_other"],
        "cos_other_actual": actuals["cos_other"],
        "cos_other_var": actuals["cos_other"] - targets["cos_other"],
        "rsb_paint_target": targets["rsb_paint"],
        "rsb_paint_actual": actuals["rsb_paint"],
        "rsb_paint_var": actuals["rsb_paint"] - targets["rsb_paint"],
        "consumables_target": targets["consumables"],
        "consumables_actual": actuals["consumables_combined"],
        "consumables_var": actuals["consumables_combined"] - targets["consumables"],
        "gp_target": targets["gross_profit"],
        "gp_actual": actuals["gross_profit"],
        "gp_var": actuals["gross_profit"] - targets["gross_profit"],
        "gp_pct_target": targets["gp_pct"] / 100,  # store as decimal for Excel %
        "gp_pct_actual": actuals["gp_pct"] / 100,
        "diagnostics_target": extras.get("diagnostics_target", 0),
        "diagnostics_actual": extras.get("diagnostics", 0),
        "additionals_target": extras.get("additionals_target", 0),
        "additionals_actual": extras.get("additionals", 0),
        "csi_target": extras.get("csi_target", 0) / 100,  # decimal for Excel %
        "csi_actual": extras.get("csi", 0) / 100,
    }


def _total_row(rows: list[dict]) -> dict:
    """Build the TOTAL aggregation row for the Summary table.

    Sums for currency columns. Weighted GP% = ΣGP / ΣSales.
    Weighted CSI = Σ(CSI × Sales) / ΣSales (uses actual sales as weight).
    """
    sum_keys = [
        "sales_target", "sales_actual", "sales_var",
        "parts_costs_target", "parts_costs_actual", "parts_costs_var",
        "cos_other_target", "cos_other_actual", "cos_other_var",
        "rsb_paint_target", "rsb_paint_actual", "rsb_paint_var",
        "consumables_target", "consumables_actual", "consumables_var",
        "gp_target", "gp_actual", "gp_var",
        "diagnostics_target", "diagnostics_actual",
        "additionals_target", "additionals_actual",
    ]
    totals = {"branch": "TOTAL"}
    for k in sum_keys:
        totals[k] = sum(r[k] for r in rows)

    sales_t = totals["sales_target"]
    sales_a = totals["sales_actual"]
    totals["gp_pct_target"] = (totals["gp_target"] / sales_t) if sales_t > 0 else 0.0
    totals["gp_pct_actual"] = (totals["gp_actual"] / sales_a) if sales_a > 0 else 0.0

    weighted_csi_t = sum(r["csi_target"] * r["sales_target"] for r in rows)
    weighted_csi_a = sum(r["csi_actual"] * r["sales_actual"] for r in rows)
    totals["csi_target"] = (weighted_csi_t / sales_t) if sales_t > 0 else 0.0
    totals["csi_actual"] = (weighted_csi_a / sales_a) if sales_a > 0 else 0.0

    return totals


def build_month_report(
    month: str,
    cos_other_pct: float,
    budget_df,
    actuals_df,
) -> dict:
    """Assemble the cross-branch report for one month.

    Args:
        month: e.g. "March".
        cos_other_pct: configured Cost of Sales Other rate as a decimal (e.g. 0.07).
        budget_df: pandas.DataFrame returned by sheets.get_budget_data().
        actuals_df: pandas.DataFrame returned by sheets.get_actuals_data().

    Returns:
        {
            "month": str,
            "generated_at": ISO timestamp str,
            "summary_rows": [<branch row>, ..., <TOTAL row>],
            "branch_details": {
                "<branch>": {
                    "targets": dict,       # always populated (zero-filled if no budget)
                    "actuals": dict,       # always populated (zero-filled if no actuals)
                    "extras": dict,        # diagnostics/additionals/csi (target & actual)
                    "has_budget": bool,
                    "has_actuals": bool,
                },
                ...
            },
            "branches_with_budget": int,
            "branches_with_actuals": int,
        }
    """
    branch_details = {}
    summary_rows = []
    branches_with_budget = 0
    branches_with_actuals = 0

    for branch in BRANCHES:
        budget_row = None
        if budget_df is not None and not budget_df.empty:
            match = budget_df[
                (budget_df["branch"] == branch) & (budget_df["month"] == month)
            ]
            if not match.empty:
                budget_row = match.iloc[0].to_dict()

        actuals_row = None
        if actuals_df is not None and not actuals_df.empty:
            match = actuals_df[
                (actuals_df["branch"] == branch) & (actuals_df["month"] == month)
            ]
            if not match.empty:
                actuals_row = match.iloc[0].to_dict()

        if budget_row:
            targets = targets_from_budget(budget_row, cos_other_pct)
            branches_with_budget += 1
        else:
            targets = _empty_targets()

        if actuals_row:
            actuals = actuals_from_row(actuals_row, cos_other_pct)
            branches_with_actuals += 1
        else:
            actuals = calculate_gp_actuals(0, 0, 0, 0, 0, 0, cos_other_pct=cos_other_pct)

        extras = {
            "diagnostics": float(actuals_row.get("diagnostics", 0)) if actuals_row else 0.0,
            "additionals": float(actuals_row.get("additionals", 0)) if actuals_row else 0.0,
            "csi": float(actuals_row.get("csi", 0)) if actuals_row else 0.0,
            "diagnostics_target": float(budget_row.get("diagnostics_target", 0)) if budget_row else 0.0,
            "additionals_target": float(budget_row.get("additionals_target", 0)) if budget_row else 0.0,
            "csi_target": float(budget_row.get("csi_target", 0)) if budget_row else 0.0,
        }

        branch_details[branch] = {
            "targets": targets,
            "actuals": actuals,
            "extras": extras,
            "has_budget": budget_row is not None,
            "has_actuals": actuals_row is not None,
        }

        summary_rows.append(_row_for_branch(branch, targets, actuals, extras))

    summary_rows.append(_total_row(summary_rows))

    return {
        "month": month,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "summary_rows": summary_rows,
        "branch_details": branch_details,
        "branches_with_budget": branches_with_budget,
        "branches_with_actuals": branches_with_actuals,
    }


# Schema for Summary sheet — (header_label, key_in_row_dict, format_kind, is_cost)
# format_kind ∈ {"currency", "percent", "text"}
# is_cost: True means lower-is-better (variance colours invert)
_SUMMARY_COLUMNS = [
    ("Branch",                "branch",              "text",     False),
    ("Sales Target",          "sales_target",        "currency", False),
    ("Sales Actual",          "sales_actual",        "currency", False),
    ("Sales Var",             "sales_var",           "currency", False),
    ("Parts Costs Target",    "parts_costs_target",  "currency", True),
    ("Parts Costs Actual",    "parts_costs_actual",  "currency", True),
    ("Parts Costs Var",       "parts_costs_var",     "currency", True),
    ("COS Other Target",      "cos_other_target",    "currency", True),
    ("COS Other Actual",      "cos_other_actual",    "currency", True),
    ("COS Other Var",         "cos_other_var",       "currency", True),
    ("RSB Paint Target",      "rsb_paint_target",    "currency", True),
    ("RSB Paint Actual",      "rsb_paint_actual",    "currency", True),
    ("RSB Paint Var",         "rsb_paint_var",       "currency", True),
    ("Consumables Target",    "consumables_target",  "currency", True),
    ("Consumables Actual",    "consumables_actual",  "currency", True),
    ("Consumables Var",       "consumables_var",     "currency", True),
    ("GP Target",             "gp_target",           "currency", False),
    ("GP Actual",             "gp_actual",           "currency", False),
    ("GP Var",                "gp_var",              "currency", False),
    ("GP% Target",            "gp_pct_target",       "percent",  False),
    ("GP% Actual",            "gp_pct_actual",       "percent",  False),
    ("Diagnostics Target",    "diagnostics_target",  "currency", False),
    ("Diagnostics Actual",    "diagnostics_actual",  "currency", False),
    ("Additionals Target",    "additionals_target",  "currency", False),
    ("Additionals Actual",    "additionals_actual",  "currency", False),
    ("CSI Target",            "csi_target",          "percent",  False),
    ("CSI Actual",            "csi_actual",          "percent",  False),
]


def _apply_variance_colour(ws, col_letter: str, first_row: int, last_row: int, is_cost: bool):
    """Add conditional formatting: green for 'good' variance, red for 'bad'.

    For revenue columns: positive var = green, negative = red.
    For cost columns (is_cost=True): negative var = green, positive = red.
    """
    rng = f"{col_letter}{first_row}:{col_letter}{last_row}"
    if is_cost:
        good = CellIsRule(operator="lessThan", formula=["0"], fill=GREEN_FILL)
        bad = CellIsRule(operator="greaterThan", formula=["0"], fill=RED_FILL)
    else:
        good = CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL)
        bad = CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL)
    ws.conditional_formatting.add(rng, good)
    ws.conditional_formatting.add(rng, bad)


def _write_summary_sheet(wb: Workbook, report: dict):
    """Populate the 'Summary' sheet with one row per branch + TOTAL."""
    ws = wb.active
    ws.title = "Summary"

    # Title row
    ws.cell(row=1, column=1, value=f"Cross-Branch Summary — {report['month']}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated {report['generated_at']}").font = Font(italic=True, color="666666")

    header_row = 4
    first_data_row = 5

    # Headers
    for col_idx, (label, _key, _fmt, _is_cost) in enumerate(_SUMMARY_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    rows = report["summary_rows"]
    for row_offset, row in enumerate(rows):
        r = first_data_row + row_offset
        is_total = (row.get("branch") == "TOTAL")
        for col_idx, (_label, key, fmt, _is_cost) in enumerate(_SUMMARY_COLUMNS, start=1):
            cell = ws.cell(row=r, column=col_idx, value=row.get(key, 0))
            if fmt == "currency":
                cell.number_format = CURRENCY_FMT
            elif fmt == "percent":
                cell.number_format = PCT_FMT
            if is_total:
                cell.font = TOTAL_FONT
                cell.fill = SECTION_FILL

    # Conditional formatting on variance columns (data rows only — not TOTAL,
    # since totals already convey the aggregate)
    last_branch_row = first_data_row + len(rows) - 2  # exclude TOTAL row from CF
    for col_idx, (_label, key, _fmt, is_cost) in enumerate(_SUMMARY_COLUMNS, start=1):
        if not key.endswith("_var"):
            continue
        col_letter = get_column_letter(col_idx)
        _apply_variance_colour(ws, col_letter, first_data_row, last_branch_row, is_cost)

    # Column widths
    ws.column_dimensions["A"].width = 24
    for col_idx in range(2, len(_SUMMARY_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    # Freeze panes: top header rows + first column
    ws.freeze_panes = "B5"


def _write_branch_sheet(wb: Workbook, branch: str, detail: dict, month: str, generated_at: str):
    """Write one sheet for a single branch — mirrors the dashboard GP table layout."""
    ws = wb.create_sheet(title=branch[:31])  # Excel sheet-name limit
    targets = detail["targets"]
    actuals = detail["actuals"]
    extras = detail["extras"]

    # Caption
    ws.cell(row=1, column=1, value=f"{branch} — {month} — Generated {generated_at}").font = Font(italic=True, color="666666")

    # Compute Act % values (as decimals for PCT_FMT)
    a_sales = actuals["sales"]
    def pct_of_sales(actual_value: float) -> float:
        return (actual_value / a_sales) if a_sales > 0 else 0.0

    t_sales = targets["sales_target"]
    def t_pct_of_sales(target_value: float) -> float:
        return (target_value / t_sales) if t_sales > 0 else 0.0

    # Section: Targets
    ws.cell(row=3, column=1, value="Targets").font = SECTION_FONT
    ws.cell(row=3, column=1).fill = SECTION_FILL

    ws.cell(row=4, column=1, value="Sales Target")
    ws.cell(row=4, column=2, value=targets["sales_target"]).number_format = CURRENCY_FMT

    ws.cell(row=5, column=1, value="Paint Labour Other Target")
    ws.cell(row=5, column=2, value=targets["paint_labour_target"]).number_format = CURRENCY_FMT

    ws.cell(row=6, column=1, value="Parts Sales Target")
    ws.cell(row=6, column=2, value=targets["parts_sales_target"]).number_format = CURRENCY_FMT

    ws.cell(row=7, column=1, value="Parts Markup")
    ws.cell(row=7, column=2, value=targets["parts_markup"] / 100).number_format = PCT_FMT

    # Section: GP Calculations
    ws.cell(row=9, column=1, value="GP Calculations").font = SECTION_FONT
    ws.cell(row=9, column=1).fill = SECTION_FILL

    headers = ["Item", "Target", "Actual", "Variance", "Act %", "Target %"]
    for col_idx, label in enumerate(headers, start=1):
        c = ws.cell(row=10, column=col_idx, value=label)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Rows: (label, target, actual, is_cost)
    gp_rows = [
        ("Sales",                targets["sales_target"], actuals["sales"], False),
        ("Parts Costs",          targets["parts_costs"],  actuals["parts_costs"], True),
        ("Cost of Sales Other",  targets["cos_other"],    actuals["cos_other"], True),
        ("RSB Paint",            targets["rsb_paint"],    actuals["rsb_paint"], True),
        ("Consumables Combined", targets["consumables"],  actuals["consumables_combined"], True),
    ]
    var_first_row = 11
    for offset, (label, t_val, a_val, is_cost) in enumerate(gp_rows):
        r = var_first_row + offset
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=t_val).number_format = CURRENCY_FMT
        ws.cell(row=r, column=3, value=a_val).number_format = CURRENCY_FMT
        ws.cell(row=r, column=4, value=a_val - t_val).number_format = CURRENCY_FMT
        ws.cell(row=r, column=5, value=pct_of_sales(a_val)).number_format = PCT_FMT
        ws.cell(row=r, column=6, value=t_pct_of_sales(t_val)).number_format = PCT_FMT
        # Per-row variance colouring via conditional formatting
        col_letter = get_column_letter(4)
        _apply_variance_colour(ws, col_letter, r, r, is_cost)

    var_last_row = var_first_row + len(gp_rows) - 1
    gp_row = var_last_row + 1
    pct_row = var_last_row + 2

    ws.cell(row=gp_row, column=1, value="Gross Profit").font = TOTAL_FONT
    ws.cell(row=gp_row, column=2, value=targets["gross_profit"]).number_format = CURRENCY_FMT
    ws.cell(row=gp_row, column=3, value=actuals["gross_profit"]).number_format = CURRENCY_FMT
    ws.cell(row=gp_row, column=4, value=actuals["gross_profit"] - targets["gross_profit"]).number_format = CURRENCY_FMT
    for c in range(1, 5):
        ws.cell(row=gp_row, column=c).font = TOTAL_FONT
    _apply_variance_colour(ws, "D", gp_row, gp_row, False)

    ws.cell(row=pct_row, column=1, value="GP %").font = TOTAL_FONT
    ws.cell(row=pct_row, column=2, value=targets["gp_pct"] / 100).number_format = PCT_FMT
    ws.cell(row=pct_row, column=3, value=actuals["gp_pct"] / 100).number_format = PCT_FMT
    ws.cell(row=pct_row, column=4, value=(actuals["gp_pct"] - targets["gp_pct"]) / 100).number_format = PCT_FMT
    for c in range(1, 5):
        ws.cell(row=pct_row, column=c).font = TOTAL_FONT

    # Section: Additional KPIs
    kpi_section_row = pct_row + 2
    ws.cell(row=kpi_section_row, column=1, value="Additional KPIs").font = SECTION_FONT
    ws.cell(row=kpi_section_row, column=1).fill = SECTION_FILL

    kpi_rows = [
        ("Diagnostics", extras.get("diagnostics_target", 0), extras.get("diagnostics", 0), "currency"),
        ("Additionals", extras.get("additionals_target", 0), extras.get("additionals", 0), "currency"),
        ("CSI",         extras.get("csi_target", 0) / 100,   extras.get("csi", 0) / 100,   "percent"),
    ]
    for offset, (label, t_val, a_val, fmt) in enumerate(kpi_rows):
        r = kpi_section_row + 1 + offset
        ws.cell(row=r, column=1, value=label)
        cell_t = ws.cell(row=r, column=2, value=t_val)
        cell_a = ws.cell(row=r, column=3, value=a_val)
        if fmt == "currency":
            cell_t.number_format = CURRENCY_FMT
            cell_a.number_format = CURRENCY_FMT
            ws.cell(row=r, column=4, value=a_val - t_val).number_format = CURRENCY_FMT
        else:
            cell_t.number_format = PCT_FMT
            cell_a.number_format = PCT_FMT
            # No variance cell for CSI (matches dashboard behaviour)

    # Column widths
    ws.column_dimensions["A"].width = 28
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 16

    # Freeze top 2 rows
    ws.freeze_panes = "A3"


def build_excel_workbook(report: dict) -> bytes:
    """Serialize a report dict (from build_month_report) to xlsx bytes."""
    wb = Workbook()
    _write_summary_sheet(wb, report)
    for branch, detail in report["branch_details"].items():
        _write_branch_sheet(wb, branch, detail, report["month"], report["generated_at"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
